import tkinter as tk
from tkinter import ttk
from openpyxl import *
from tkinter import messagebox

root = tk.Tk()
root.title("Hotel hungry")

root.iconbitmap(r"C:\Users\Abhinav\Downloads\biryani_ExO_icon.ico")

wb = load_workbook(r"C:\Users\Abhinav\Desktop\Python\rmapp.xlsx")
sheet = wb.active

starter = ["< select your starters >", "None", "Veg Manchurian", "Tandoori", "Chicken 65", "Chicken Manchurian",
           "Gobi 65", "Noodles", "Samosa", "Chaat", "Pani Puri"]
main_c = ["< select your main course >", "None", "Veg Biryani", "Chicken Biryani", "Roti", "Jeera Rice", "Fried Rice",
          "Naan", "Phulka"]
drinks = ["< select your drinks >", "None", "Coca cola", "Sprite", "Maaza", "Thumps Up", "Fanta", "Pepsi",
          " Virgin Mohito"]
deserts = ["< select your desert >", "None", "Fruit box", "Chocolate Ice cream", "Vannila Ice cream", "Pista Icecream",
           "Butterskotch icecream", "Brownie", "Casata"]

name_l = tk.Label(root, text="Name: ")
name_l.config(font=("Comic Sans MS", 12))
name_l.grid(row=1, column=1, pady=2)

starters_l = tk.Label(root, text="Select Starters: ")
starters_l.config(font=("Comic Sans MS", 12))
starters_l.grid(row=2, column=1, pady=2)

q1 = tk.Label(root, text=" - Quantity (in plates): ")
q1.config(font=("Comic Sans MS", 12))
q1.grid(row=2, column=3, pady=2)

mc_l = tk.Label(root, text="Select Main Course: ")
mc_l.config(font=("Comic Sans MS", 12))
mc_l.grid(row=3, column=1, pady=2)

q2 = tk.Label(root, text=" - Quantity (in plates): ")
q2.config(font=("Comic Sans MS", 12))
q2.grid(row=3, column=3, pady=2)

drink_l = tk.Label(root, text="Select Drinks: ")
drink_l.config(font=("Comic Sans MS", 12))
drink_l.grid(row=4, column=1, pady=2)

q3 = tk.Label(root, text=" - Quantity (in bottle - 250 ml): ")
q3.config(font=("Comic Sans MS", 12))
q3.grid(row=4, column=3, pady=2)

deserts_l = tk.Label(root, text="Select Deserts: ")
deserts_l.config(font=("Comic Sans MS", 12))
deserts_l.grid(row=5, column=1, pady=2)

q4 = tk.Label(root, text=" - Quantity (in bowls): ")
q4.config(font=("Comic Sans MS", 12))
q4.grid(row=5, column=3, pady=2)

tip_l = tk.Label(root, text="Enter your tip: ")
tip_l.config(font=("Comic Sans MS", 12))
tip_l.grid(row=6, column=1, pady=2)

var1 = tk.StringVar()
name_e = tk.Entry(root, textvariable=var1)
name_e.grid(row=1, column=2)

var2 = tk.StringVar()
s_cb = ttk.Combobox(root, width=27, textvariable=var2)
s_cb["values"] = starter
s_cb.current(0)
s_cb.grid(row=2, column=2)

var3 = tk.StringVar()
mc_cb = ttk.Combobox(root, width=27, textvariable=var3)
mc_cb["values"] = main_c
mc_cb.current(0)
mc_cb.grid(row=3, column=2)

var4 = tk.StringVar()
d_cb = ttk.Combobox(root, width=27, textvariable=var4)
d_cb["values"] = drinks
d_cb.current(0)
d_cb.grid(row=4, column=2)

var5 = tk.StringVar()
de_cb = ttk.Combobox(root, width=27, textvariable=var5)
de_cb["values"] = deserts
de_cb.current(0)
de_cb.grid(row=5, column=2)

var6 = tk.StringVar()
t_cb = tk.Entry(root, textvariable=var6)
t_cb.grid(row=6, column=2)

var7 = tk.StringVar()
q1e = tk.Entry(root, textvariable=var7)
q1e.grid(row=2, column=4)

var8 = tk.StringVar()
q2e = tk.Entry(root, textvariable=var8)
q2e.grid(row=3, column=4)

var9 = tk.StringVar()
q3e = tk.Entry(root, textvariable=var9)
q3e.grid(row=4, column=4)

var10 = tk.StringVar()
q4e = tk.Entry(root, textvariable=var10)
q4e.grid(row=5, column=4)


def submit():
    n_g = var1.get()
    s_g = var2.get()
    mc_g = var3.get()
    d_g = var4.get()
    de_g = var5.get()
    t_g = var6.get()

    g1 = q1e.get()
    g2 = q2e.get()
    g3 = q3e.get()
    g4 = q4e.get()

    cur_row = sheet.max_row

    sheet.cell(row=cur_row + 1, column=1).value = n_g
    sheet.cell(row=cur_row + 1, column=2).value = s_g
    sheet.cell(row=cur_row + 1, column=3).value = int(g1)
    sheet.cell(row=cur_row + 1, column=4).value = mc_g
    sheet.cell(row=cur_row + 1, column=5).value = int(g3)
    sheet.cell(row=cur_row + 1, column=6).value = d_g
    sheet.cell(row=cur_row + 1, column=7).value = int(g4)
    sheet.cell(row=cur_row + 1, column=8).value = de_g
    sheet.cell(row=cur_row + 1, column=9).value = t_g

    wb.save(r"C:\Users\Abhinav\Desktop\Python\rmapp.xlsx")
    name_e.focus_set()
    name_e.delete(0, tk.END)

    s_cb.focus_set()
    s_cb.delete(0, tk.END)

    mc_cb.focus_set()
    mc_cb.delete(0, tk.END)

    d_cb.focus_set()
    d_cb.delete(0, tk.END)

    de_cb.focus_set()
    de_cb.delete(0, tk.END)

    t_cb.focus_set()
    t_cb.delete(0, tk.END)

    q1e.focus_set()
    q1e.delete(0, tk.END)

    q2e.focus_set()
    q2e.delete(0, tk.END)

    q3e.focus_set()
    q3e.delete(0, tk.END)

    q4e.focus_set()
    q4e.delete(0, tk.END)

    messagebox.showinfo("message from hotel hungry",
                        message=" our client will call you in 15 min \n Thankyou for choosing Hotel Hungry!! ")


btn = tk.Button(root, text="SUBMIT", command=submit)
btn.config(font=("Comic Sans MS", 12))
btn.grid(row=7, column=2)

root.mainloop()